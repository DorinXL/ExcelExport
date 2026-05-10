from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import (
    DATA_START_ROW_INDEX,
    HEADER_ROW_INDEX,
    TYPE_ROW_INDEX,
    DEFAULT_VALUE_ROW_INDEX,
    COMMENT_ROW_INDEX,
    SKIP_SHEET_PREFIXES,
    TYPE_MAP,
)
from utils import is_blank, is_blank_row, normalize_text, trim_trailing_empty_columns


@dataclass
class SheetData:
    file_name: str
    sheet_name: str
    headers: list[str]
    types: list[str]
    comments: list[str]
    records: list[dict[str, Any]]


def parse_excel_file(filepath: Path) -> tuple[list[SheetData], list[str]]:
    sheets: list[SheetData] = []
    errors: list[str] = []

    try:
        workbook = load_workbook(filepath, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl error details depend on file content
        return [], [f"{filepath.name}：无法读取文件，原因：{exc}"]

    try:
        for worksheet in workbook.worksheets:
            if should_skip_sheet(worksheet.title):
                continue

            sheet_data, sheet_errors = parse_sheet(worksheet, filepath.name)
            errors.extend(sheet_errors)
            if sheet_data is not None:
                sheets.append(sheet_data)
    finally:
        workbook.close()

    return sheets, errors


def should_skip_sheet(sheet_name: str) -> bool:
    return sheet_name.startswith(SKIP_SHEET_PREFIXES)


def parse_sheet(worksheet: Any, file_name: str) -> tuple[SheetData | None, list[str]]:
    errors: list[str] = []
    sheet_name = worksheet.title

    if worksheet.max_row < DATA_START_ROW_INDEX:
        return None, [
            f"{file_name} - Sheet '{sheet_name}'：表头行数不足（至少需要5行：名称、类型、默认值、英文标识、中文标识）"
        ]

    types = [
        normalize_text(worksheet.cell(row=TYPE_ROW_INDEX + 1, column=col).value)
        for col in range(1, worksheet.max_column + 1)
    ]
    default_values = [
        normalize_text(worksheet.cell(row=DEFAULT_VALUE_ROW_INDEX + 1, column=col).value)
        for col in range(1, worksheet.max_column + 1)
    ]
    headers = [
        normalize_text(worksheet.cell(row=HEADER_ROW_INDEX + 1, column=col).value)
        for col in range(1, worksheet.max_column + 1)
    ]
    comments = [
        normalize_text(worksheet.cell(row=COMMENT_ROW_INDEX + 1, column=col).value)
        for col in range(1, worksheet.max_column + 1)
    ]
    comments, headers, types, default_values = trim_trailing_empty_columns(comments, headers, types, default_values)

    if not headers:
        return None, [f"{file_name} - Sheet '{sheet_name}'：字段名为空，无法导出"]

    for index, header in enumerate(headers):
        column_name = header or f"第 {index + 1} 列"
        type_mark = types[index] if index < len(types) else ""
        if header.strip() == "":
            errors.append(f"{file_name} - Sheet '{sheet_name}'：第 {index + 1} 列字段名为空")
        if type_mark.strip() == "":
            errors.append(f"{file_name} - Sheet '{sheet_name}' - 字段 '{column_name}'：缺少类型标记")
        elif type_mark not in TYPE_MAP:
            errors.append(
                f"{file_name} - Sheet '{sheet_name}' - 字段 '{column_name}'：未知的类型标记 '{type_mark}'"
            )

    if errors:
        return None, errors

    records: list[dict[str, Any]] = []
    has_data_row = False

    for row_index in range(DATA_START_ROW_INDEX + 1, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_index, column=col).value for col in range(1, len(headers) + 1)
        ]
        
        # Check for # prefix to skip row
        if row_values and not is_blank(row_values[0]) and str(row_values[0]).strip().startswith("#"):
            continue

        if is_blank_row(row_values):
            continue

        has_data_row = True
        record, row_errors = parse_data_row(
            row_values=row_values,
            row_index=row_index,
            headers=headers,
            types=types,
            default_values=default_values,
            file_name=file_name,
            sheet_name=sheet_name,
        )
        errors.extend(row_errors)
        if record is not None:
            records.append(record)

    if not has_data_row:
        return None, []

    return (
        SheetData(
            file_name=file_name,
            sheet_name=sheet_name,
            headers=headers,
            types=types,
            comments=comments,
            records=records,
        ),
        errors,
    )


def parse_data_row(
    row_values: list[Any],
    row_index: int,
    headers: list[str],
    types: list[str],
    default_values: list[str],
    file_name: str,
    sheet_name: str,
) -> tuple[dict[str, Any] | None, list[str]]:
    record: dict[str, Any] = {}
    errors: list[str] = []

    for header, type_mark, default_value, raw_value in zip(headers, types, default_values, row_values):
        if is_blank(raw_value) and not is_blank(default_value):
            raw_value = default_value

        parsed_value, error = parse_value(raw_value, type_mark)
        if error is not None:
            errors.append(
                f"{file_name} - Sheet '{sheet_name}' - 第 {row_index} 行 '{header}' 列：{error}"
            )
        record[header] = parsed_value

    return record, errors


def parse_value(raw_value: Any, type_mark: str) -> tuple[Any, str | None]:
    if type_mark == "string":
        if raw_value is None:
            return "", None
        return str(raw_value), None

    if type_mark == "bool":
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            return False, None
        if isinstance(raw_value, bool):
            return raw_value, None
        if isinstance(raw_value, (int, float)):
            return raw_value != 0, None
        return str(raw_value).strip().lower() in {"true", "1"}, None

    if type_mark == "int":
        return parse_int_value(raw_value)

    if type_mark == "float":
        return parse_float_value(raw_value)

    if type_mark == "array_int":
        return parse_array_value(raw_value, parse_int_value)

    if type_mark == "array_float":
        return parse_array_value(raw_value, parse_float_value)

    if type_mark == "array_string":
        if is_blank(raw_value):
            return [], None
        text = str(raw_value).strip()
        if not (text.startswith("[") and text.endswith("]")):
            return None, f"期望数组使用 [...] 格式，实际值为 '{raw_value}'"
        inner_text = text[1:-1].strip()
        if not inner_text:
            return [], None
        return [part.strip() for part in inner_text.split(",") if part.strip() != ""], None

    return None, f"未知类型标记 '{type_mark}'"


def parse_int_value(raw_value: Any) -> tuple[int | None, str | None]:
    if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
        return 0, None

    try:
        if isinstance(raw_value, bool):
            return int(raw_value), None
        if isinstance(raw_value, int):
            return raw_value, None
        if isinstance(raw_value, float):
            if raw_value.is_integer():
                return int(raw_value), None
            raise ValueError

        text = str(raw_value).strip()
        if "." in text:
            number = float(text)
            if not number.is_integer():
                raise ValueError
            return int(number), None
        return int(text), None
    except (TypeError, ValueError):
        return None, f"期望 int 类型，实际值为 '{raw_value}'"


def parse_float_value(raw_value: Any) -> tuple[float | None, str | None]:
    if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
        return 0.0, None

    try:
        return float(raw_value), None
    except (TypeError, ValueError):
        return None, f"期望 float 类型，实际值为 '{raw_value}'"


def parse_array_value(
    raw_value: Any, item_parser: Any
) -> tuple[list[Any] | None, str | None]:
    if is_blank(raw_value):
        return [], None

    text = str(raw_value).strip()
    if not (text.startswith("[") and text.endswith("]")):
        return None, f"期望数组使用 [...] 格式，实际值为 '{raw_value}'"

    inner_text = text[1:-1].strip()
    if not inner_text:
        return [], None

    results: list[Any] = []
    parts = inner_text.split(",")
    for part in parts:
        stripped = part.strip()
        if stripped == "":
            continue
        parsed, error = item_parser(stripped)
        if error is not None:
            return None, f"期望数组元素类型正确，实际值为 '{raw_value}'"
        results.append(parsed)
    return results, None
